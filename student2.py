import os
import re
from matplotlib import pyplot as plt
import openpyxl
import xlrd
import pandas as pd
class Student:                       #定义一个学生类
 def __init__(self):
    self.name = ''
    self.ID =''
    self.score1 = 0
    self.score2 = 0
    self.score3 = 0
    self.sum = 0
 def sumscore(self):                #计算总分
     self.sum=self.score1 + self.score2 + self.score3
 
def searchByID(stulist, ID):         #按学号查找看是否学号已经存在
  for item in stulist:
     if item.ID == ID:
         return True
 
def Add(stulist,stu):                #添加一个学生信息
     if searchByID(stulist, stu.ID) == True:
         print("学号已经存在！")
         return False
     stulist.append(stu)
     print(stu.name,stu.ID, stu.score1, stu.score2, stu.score3, stu.sum)
     print("是否要保存学生信息？")
     nChoose = input("Choose Y/N")
     if nChoose == 'Y' or nChoose == 'y':
         file_object = open("students.txt", "a")
         file_object.write(stu.ID)
         file_object.write(" ")
         file_object.write(stu.name)
         file_object.write(" ")
         file_object.write(str(stu.score1))
         file_object.write(" ")
         file_object.write(str(stu.score2))
         file_object.write(" ")
         file_object.write(str(stu.score3))
         file_object.write(" ")
         file_object.write(str(stu.sum))
         file_object.write("\n")
         file_object.close()
         print("保存成功！")
 
def Search(stulist, ID):             #搜索一个学生信息
 print("学号\t姓名\t语文\t数学\t英语\t总分")
 count = 0
 for item in stulist:
     if item.ID == ID:
         print(item.ID, '\t' ,item.name,'\t', item.score1,'\t',item.score2, '\t', item.score3, '\t',item.sum)
         break
     count = count + 1
 if count == len(stulist):
     print("没有该学生学号！")
 
def Del(stulist, ID):                #删除一个学生信息
 count = 0
 for item in stulist:
      if item.ID == ID:
           stulist.remove(item)
           print("删除成功！")
           break
      count +=1
 if count == len(stulist):
     print("没有该学生学号！")
 file_object = open("students.txt", "w",encoding="utf-8")  #覆盖写入
 for stu in stulist:
     print (stu.ID, stu.name, stu.score1,stu.score2, stu.score3, stu.sum)
     file_object.write(stu.ID)
     file_object.write(" ")
     file_object.write(stu.name)
     file_object.write(" ")
     file_object.write(str(stu.score1))
     file_object.write(" ")
     file_object.write(str(stu.score2))
     file_object.write(" ")
     file_object.write(str(stu.score3))
     file_object.write(" ")
     file_object.write(str(stu.sum))
     file_object.write("\n")
 # print("删除保存成功！")
 file_object.close()
     
def Change(stulist, ID):             #修改学生信息
  count = 0
  for item in stulist:
     if item.ID == ID:
        stulist.remove(item)
        file_object = open("students.txt", "w")
        for stu in stulist:
          #print li.ID, li.name, li.score
          file_object.write(stu.ID)
          file_object.write(" ")
          file_object.write(stu.name)
          file_object.write(" ")
          file_object.write(str(stu.score1))
          file_object.write(" ")
          file_object.write(str(stu.score2))
          file_object.write(" ")
          file_object.write(str(stu.score3))
          file_object.write(" ")
          file_object.write(str(stu.sum))
          file_object.write("\n")
        file_object.close()
  #输入这个被修改学生的新信息
  stu = Student()
  stu.name = input("请输入学生的姓名")
  stu.ID = input("请输入学生的ID")
  stu.score1 = int(input("请输入学生语文成绩"))
  stu.score2 = int(input("请输入学生数学成绩"))
  stu.score3 = int(input("请输入学生英语成绩"))
  stu.sumscore()
  Add(stulist,stu) #添加一个stu学生信息到文件中
             
def display(stulist):                #显示所有学生信息
    print("学号\t姓名  语文 数学 英语 总分")
    for item in stulist:
        #print(item.ID, '\t' ,item.name,'\t', item.score1,'\t',item.score2, '\t', item.score3, '\t',item.sum)
        print("%5s %5s %3d  %3d  %3d  %4d"%(item.ID,item.name,item.score1,item.score2,item.score3,item.sum))
        
def Sort(stulist):                   #按学生成绩排序
     insertSort(stulist)     
     display(stulist)    
def insertSort(stulist):           
 for i in range(len(stulist)-1): 
    for j in range(i+1,len(stulist)): 
        if stulist[i].sum<stulist[j].sum:
            temp = stulist[i] 
            stulist[i] = stulist[j] 
            stulist[j] = temp         

def read_students():                 #txt文件转文xlsx文件
    with open('students.txt', 'r') as f:
        data = f.readlines()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "学号"
    sheet.cell(row=1, column=2).value = "姓名"
    sheet.cell(row=1, column=3).value = "语文成绩"
    sheet.cell(row=1, column=4).value = "数学成绩"
    sheet.cell(row=1, column=5).value = "英语成绩"
    sheet.cell(row=1, column=6).value = "总分"
    for i in range(len(data)):
        row_data = re.split('\t|\s+', data[i].strip())
        sheet.cell(row=i+2, column=1).value = row_data[0]
        sheet.cell(row=i+2, column=2).value = row_data[1]
        for j in range(len(row_data)-2):
            sheet.cell(row=i+2, column=j+3).value = row_data[j+2]
        if row_data[-1].isdigit():
            sheet.cell(row=i+2, column=6).value = int(row_data[-1])
    wb.save('data.xlsx')
    df = pd.read_excel('data.xlsx')
    return df
    
def show_scores():                   #成绩柱状图展示
    try:
        # 加载数据
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.active
        header = [cell.value for cell in sheet[1]][2:]
        
        # 选择课程
        print(f"当前支持的课程为：{','.join(header)}")
        course_names = input("请输入需要展示的课程名，多个课程以逗号分隔：")
        selected_courses = course_names.split(',')
        for course in selected_courses:
            if course not in header:
                raise ValueError(f"输入课程“{course}”错误！")
        course_cols = [header.index(c)+2 for c in selected_courses]
        df_data = [[row[0].value, row[1].value]+[row[c].value for c in course_cols] for row in sheet.iter_rows(min_row=2, values_only=False)]
        df = pd.DataFrame(df_data, columns=["姓名", "学号"]+selected_courses)
        df = df.set_index("姓名")
        for course in selected_courses:
            df[course] = pd.to_numeric(df[course], errors='coerce')

        # 处理数据
        bins = [0, 60, 70, 80, 90, 100]
        plt.figure(figsize=(8, 6), dpi=80)
        plt.rcParams['font.sans-serif'] = ['SimHei'] # 指定中文字体
        for i, course in enumerate(selected_courses):
            scores = df[course].dropna()
            score_counts = [0] * 5
            for score in scores:
                for j, bin in enumerate(bins):
                    if score < bin:
                        score_counts[j-1] += 1
                        break
                else:
                    score_counts[-1] += 1
            plt.subplot(len(selected_courses), 1, i+1)
            plt.bar(['<60分', '60~69分', '70~79分', '80~89分', '>=90分'], score_counts, color='green')
            plt.title(f"{course}成绩分布")
            plt.xlabel("分数段")
            plt.ylabel("人数")
            for j, count in enumerate(score_counts):
                x, y = j, count
                if y > 0:
                    plt.text(x, y+0.1, str(y), ha='center', va='bottom')
        plt.tight_layout()
        plt.show()
    except Exception as e:
        print(f"展示成绩柱状图时出错：{e}")
   
def Init(stulist):                   #初始化函数
     print("初始化......")
     if os.path.exists('students.txt'):          
         file_object = open('students.txt', 'r')
         for line in file_object:
             stu = Student()
             line = line.strip("\n")
             s = line.split(" ")            #按空格分隔形成列表
             stu.ID = s[0]
             stu.name = s[1]
             stu.score1 = int(s[2])
             stu.score2 = int(s[3])
             stu.score3 = int(s[4])
             stu.sum = int(s[5])
             stulist.append(stu)
         file_object.close()
         print("初始化成功！")
     main()

def main():                          #主函数 该程序的入口函数
  while True:
     print("*********************")
     print("--------菜单---------")
     print("增加学生信息--------1")
     print("查找学生信息--------2")
     print("删除学生信息--------3")
     print("修改学生信息--------4")
     print("所有学生信息--------5")
     print("按照分数排序--------6")
     print("成绩柱状图展示------7")
     print("退出程序------------0")
     print("*********************")
     
     nChoose = input("请输入你的选择：")
     if nChoose == "1":
         stu = Student()
         stu.name = input("请输入学生的姓名")
         stu.ID = input("请输入学生的ID")
         stu.score1 = int(input("请输入学生语文成绩"))
         stu.score2 = int(input("请输入学生数学成绩"))
         stu.score3 = int(input("请输入学生英语成绩"))
         stu.sumscore()
         Add(stulist,stu) 
     if nChoose == '2':
         ID = input("请输入学生的ID")
         Search(stulist, ID)
     
     if nChoose == '3':
         ID = input("请输入学生的ID")
         Del(stulist, ID)       
     if nChoose == '4':
         ID = input("请输入学生的ID")
         Change(stulist, ID)
     
     if nChoose == '5':
         display(stulist)
     
     if nChoose == '6':
         Sort(stulist) 
     if nChoose == '7':
         read_students()
         show_scores()
     if nChoose == '0':
         break
       
                             #主程序 
if __name__ == '__main__':
   stulist =[]
   Init(stulist)
   